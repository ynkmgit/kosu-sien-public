#!/usr/bin/env python3
"""実施工数管理.xlsmからASPITPJ V2A9案件の日別×担当者工数を抽出

列構造: A=セクション名, B=担当者名, C=合計, D=マーカー(head1/line1/foot1), E=行数, F〜=日別工数
行4にdatetimeで日付が入っている
"""
import json
import sys
from pathlib import Path
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).parent.parent
EXCEL_PATH = PROJECT_ROOT / "sample" / "実施工数管理.xlsm"
OUTPUT_PATH = PROJECT_ROOT / "sample" / "_worklog_aspit.json"

# DB登録済みV2A9案件CDリスト
TARGET_CDS = [
    "V2A9-415", "V2A9-870", "V2A9-1011", "V2A9-1017", "V2A9-1040",
    "V2A9-1068", "V2A9-1087", "V2A9-1094", "V2A9-1116", "V2A9-1118",
    "V2A9-1159", "V2A9-1166", "V2A9-1204", "V2A9-1215", "V2A9-1228",
    "V2A9-1238", "V2A9-1243", "V2A9-1245", "V2A9-1249", "V2A9-1250",
    "V2A9-1253", "V2A9-1261", "V2A9-1262", "V2A9-1269", "V2A9-1274",
    "V2A9-1281", "V2A9-1291", "V2A9-1292", "V2A9-1296", "V2A9-1299",
    "V2A9-1302", "V2A9-1308", "V2A9-1311", "V2A9-1315",
]


def main():
    wb = load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    ws = wb["2月"]

    all_rows = list(ws.iter_rows(min_row=1, values_only=False))

    # 行4から日付マッピング (F列=col5以降)
    date_row = all_rows[3]  # 0-indexed row 4
    date_map = {}
    for col_idx in range(5, len(date_row)):
        val = date_row[col_idx].value
        if val and hasattr(val, 'strftime'):
            date_map[col_idx] = val.strftime("%Y-%m-%d")

    # セクション解析
    result = []
    current_cd = None
    current_entries = []
    in_target = False

    for row in all_rows:
        d_val = str(row[3].value or "").strip() if len(row) > 3 else ""
        a_val = str(row[0].value or "").strip()

        if d_val == "head1":
            # 前のセクションを保存
            if in_target and current_entries:
                result.append({
                    "issue_cd": current_cd,
                    "entries": current_entries
                })

            # 新セクション開始
            # A列: "V2A9-415 案件名..."
            parts = a_val.split()
            cd = parts[0] if parts else ""
            in_target = cd in TARGET_CDS
            current_cd = cd
            current_entries = []

        elif d_val in ("line1", "line2", "line3") and in_target:
            member = str(row[1].value or "").strip()
            if not member:
                continue

            for col_idx, date_str in date_map.items():
                if col_idx < len(row):
                    val = row[col_idx].value
                    if val is not None and isinstance(val, (int, float)) and val > 0:
                        current_entries.append({
                            "user": member,
                            "date": date_str,
                            "hours": float(val)
                        })

        elif d_val == "foot1" and in_target:
            if current_entries:
                result.append({
                    "issue_cd": current_cd,
                    "entries": current_entries
                })
            in_target = False
            current_cd = None
            current_entries = []

    wb.close()

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    total_entries = sum(len(s["entries"]) for s in result)
    total_hours = sum(e["hours"] for s in result for e in s["entries"])
    print(f"V2A9 sections: {len(result)}")
    print(f"Total entries: {total_entries}")
    print(f"Total hours: {total_hours}")


if __name__ == "__main__":
    main()
